import os
import openpyxl
from openpyxl.styles import Font, Alignment

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

def create_mic20_test():
    """MIC20 format: Q1(5), Q2(5), Q3(5), Quiz(10), Assignment(5) per mid."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIC20 Test"

    # Header rows (rows 1-4)
    ws.merge_cells('C1:G1'); ws['C1'] = 'First Mid'
    ws.merge_cells('H1:L1'); ws['H1'] = 'Second Mid'
    ws['A2'] = 'Sl.No.'; ws['B2'] = 'Roll Numbers'
    ws['C2'] = 'Q1'; ws['D2'] = 'Q2'; ws['E2'] = 'Q3'; ws['F2'] = 'Quiz 1'; ws['G2'] = 'Assignment 1'
    ws['H2'] = 'Q1'; ws['I2'] = 'Q2'; ws['J2'] = 'Q3'; ws['K2'] = 'Quiz 2'; ws['L2'] = 'Assignment 2'
    ws['C3'] = 5; ws['D3'] = 5; ws['E3'] = 5; ws['F3'] = 10; ws['G3'] = 5
    ws['H3'] = 5; ws['I3'] = 5; ws['J3'] = 5; ws['K3'] = 10; ws['L3'] = 5
    ws['C4'] = 'CO 1'; ws['D4'] = 'CO 2'; ws['E4'] = 'CO 3'; ws['F4'] = 'CO 1,2,3'; ws['G4'] = 'CO 1,2,3'
    ws['H4'] = 'CO 3'; ws['I4'] = 'CO 4'; ws['J4'] = 'CO 5'; ws['K4'] = 'CO 3,4,5'; ws['L4'] = 'CO 3,4,5'

    for row in ws.iter_rows(min_row=1, max_row=4, max_col=12):
        for cell in row:
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center')

    students = [
        [1, '21H71A0301', 5, 5, 5, 10, 5,  5, 5, 5, 10, 5],
        [2, '21H71A0302', 3, 3, 3,  5, 3,  3, 3, 3,  5, 3],
        [3, '21H71A0303', 4, 2, 3,  8, 5,  3, 5, 2,  9, 5],
        [4, '21H71A0304', 1, 1, 2,  3, 1,  2, 1, 1,  4, 2],
        [5, '21H71A0305', 5, 3, 4,  9, 4,  4, 5, 3,  8, 5],
    ]

    for i, stu in enumerate(students):
        row_num = 5 + i
        for j, val in enumerate(stu):
            ws.cell(row=row_num, column=j+1, value=val)
            ws.cell(row=row_num, column=j+1).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    for col in ['C','D','E','F','G','H','I','J','K','L']:
        ws.column_dimensions[col].width = 12

    filepath = f"{OUTPUT_DIR}\\test_MIC20_input.xlsx"
    wb.save(filepath)
    print(f"MIC20 test file saved: {filepath}")
    return students


def create_mic23_test():
    """MIC23 format: paired evals per Q (10 marks each), Quiz(10), Asgn(5) per mid."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MIC23 Test"

    ws.merge_cells('D1:L1'); ws['D1'] = 'First Mid'
    ws.merge_cells('N1:V1'); ws['N1'] = 'Second Mid'
    ws['A2'] = 'Sl.No.'; ws['B2'] = 'Roll Numbers'; ws['C2'] = ''
    ws['D2'] = 'Q1 Eval1'; ws['E2'] = 'Q1 Eval2'
    ws['F2'] = 'Q2 Eval1'; ws['G2'] = 'Q2 Eval2'
    ws['H2'] = 'Q3 Eval1'; ws['I2'] = 'Q3 Eval2'
    ws['J2'] = 'Total'; ws['K2'] = 'Quiz'; ws['L2'] = 'Assignment'
    ws['M2'] = ''
    ws['N2'] = 'Q1 Eval1'; ws['O2'] = 'Q1 Eval2'
    ws['P2'] = 'Q2 Eval1'; ws['Q2'] = 'Q2 Eval2'
    ws['R2'] = 'Q3 Eval1'; ws['S2'] = 'Q3 Eval2'
    ws['T2'] = 'Total'; ws['U2'] = 'Quiz'; ws['V2'] = 'Assignment'

    ws['D3']=10; ws['E3']=10; ws['F3']=10; ws['G3']=10; ws['H3']=10; ws['I3']=10
    ws['K3']=10; ws['L3']=5
    ws['N3']=10; ws['O3']=10; ws['P3']=10; ws['Q3']=10; ws['R3']=10; ws['S3']=10
    ws['U3']=10; ws['V3']=5

    ws['D4']='CO1'; ws['E4']='CO1'; ws['F4']='CO2'; ws['G4']='CO2'; ws['H4']='CO3'; ws['I4']='CO3'
    ws['N4']='CO3'; ws['O4']='CO3'; ws['P4']='CO4'; ws['Q4']='CO4'; ws['R4']='CO5'; ws['S4']='CO5'

    for row in ws.iter_rows(min_row=1, max_row=4, max_col=22):
        for cell in row:
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center')

    # Col: Sl, Roll, blank, Q1e1,Q1e2, Q2e1,Q2e2, Q3e1,Q3e2, Total, Quiz, Asgn, blank, Q1e1,Q1e2, Q2e1,Q2e2, Q3e1,Q3e2, Total, Quiz, Asgn
    students = [
        [1, '22H75A0401', '', 10,10, 10,10, 10,10, '', 10, 5,  '', 10,10, 10,10, 10,10, '', 10, 5],
        [2, '22H75A0402', '',  8,10,  6, 9,  7, 5, '',  8, 3,  '',  7, 9,  8, 6,  5, 8, '',  7, 4],
        [3, '22H75A0403', '',  5, 5,  5, 5,  5, 5, '',  5, 3,  '',  5, 5,  5, 5,  5, 5, '',  5, 3],
        [4, '22H75A0404', '',  2, 3,  1, 4,  3, 2, '',  2, 1,  '',  3, 2,  4, 1,  2, 3, '',  3, 1],
        [5, '22H75A0405', '',  9, 7,  8, 6,  7, 9, '',  9, 5,  '',  6, 8,  9, 7,  8, 5, '',  8, 4],
    ]

    for i, stu in enumerate(students):
        row_num = 5 + i
        for j, val in enumerate(stu):
            col_num = j + 1
            ws.cell(row=row_num, column=col_num, value=val)
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    for col in ['C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V']:
        ws.column_dimensions[col].width = 10

    filepath = f"{OUTPUT_DIR}\\test_MIC23_input.xlsx"
    wb.save(filepath)
    print(f"MIC23 test file saved: {filepath}")
    return students


print("=" * 60)
print("GENERATING TEST INPUT FILES (INPUT VALUES ONLY)")
print("=" * 60)

mic20_data = create_mic20_test()
mic23_data = create_mic23_test()

print("\n" + "=" * 60)
print("MIC20 INPUT DATA")
print("=" * 60)
print(f"{'Roll':<16} {'Q1':>3} {'Q2':>3} {'Q3':>3} {'Qz1':>4} {'As1':>4}  |  {'Q1':>3} {'Q2':>3} {'Q3':>3} {'Qz2':>4} {'As2':>4}")
print("-" * 70)
for s in mic20_data:
    print(f"{s[1]:<16} {s[2]:>3} {s[3]:>3} {s[4]:>3} {s[5]:>4} {s[6]:>4}  |  {s[7]:>3} {s[8]:>3} {s[9]:>3} {s[10]:>4} {s[11]:>4}")

print("\n" + "=" * 60)
print("MIC23 INPUT DATA")
print("=" * 60)
print(f"{'Roll':<16} {'Q1e1':>4} {'Q1e2':>4} {'Q2e1':>4} {'Q2e2':>4} {'Q3e1':>4} {'Q3e2':>4} {'Qz':>3} {'As':>3}  |  {'Q1e1':>4} {'Q1e2':>4} {'Q2e1':>4} {'Q2e2':>4} {'Q3e1':>4} {'Q3e2':>4} {'Qz':>3} {'As':>3}")
print("-" * 110)
for s in mic23_data:
    print(f"{s[1]:<16} {s[3]:>4} {s[4]:>4} {s[5]:>4} {s[6]:>4} {s[7]:>4} {s[8]:>4} {s[10]:>3} {s[11]:>3}  |  {s[13]:>4} {s[14]:>4} {s[15]:>4} {s[16]:>4} {s[17]:>4} {s[18]:>4} {s[20]:>3} {s[21]:>3}")

print("\nFiles saved to OBE folder. Upload them in the app to verify!")
